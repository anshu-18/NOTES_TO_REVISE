import openpyxl as xl

wb=xl.load_workbook('stat_definitions_anap.xlsx')

print("SHEET NAMES: ", wb.sheetnames)

sheet=wb['Statistic list']
sheet1 = wb.active #If we want to work with first sheet
cell = sheet1.cell(2,1)
print(cell.value)

#printing the values inside active sheet
for val in sheet1.values:
    print(f'Row:- {val}')

#wb.save('stat_definition2.xlsx') to create a new file after making changes
